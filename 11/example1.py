from openpyxl import load_workbook

#LOAD EXCEL
wb = load_workbook(filename='11/example.xlsx')

#PRINT SHEET NAMES
print(wb.sheetnames)

#SET SHEET SPECIFIC
sheet = wb['Sheet1']

#ITER 
for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
    print(f'País: {row[0]} - Habitantes: {row[1]}') 
